import warnings
warnings.filterwarnings(action='ignore', category=UserWarning, module='gensim')
warnings.filterwarnings(action='ignore', category=FutureWarning, module='gensim')
import openpyxl
from gensim.models.word2vec import Word2Vec

pre_anti_word_list = []
after_anti_word_list = []
accent_word_list = []
trash_word_list = []

senti_word_file = openpyxl.load_workbook('C:\\Users\\inha\\Desktop\\AWRdata.xlsx')   # 장르별로 Data가 다르니 주의
sheet1 = senti_word_file.active
textlist = list()
extracted_word_list = list()

print ("리스트화 실행중")
for j in range (1,13278):                                                                    # range 직접 설정해주지 않으면 오류...해결바람
        senti_word = list()
        sentence = list()
        senti_word = sheet1.cell(row=j, column=1).value.split("+")                         # 엑셀 상의 텍스트 리스트로! 불러옵니당
        count=len(senti_word)


        for i in range (0,count):
            senti_word[i] = senti_word[i].split(",")                                       # 내부 리스트도! 불러서 2차원으로 만들어냅니다. (ㅇ)

            if (len(senti_word[i]) < 2) or (len(senti_word[i]) > 2) :
                senti_word[i] = ["error","error"]
                print("에러 텍스트 삭제")


        for k in range (0,len(senti_word)):
            if senti_word[k][1]=='Adjective':                                            # Adjective, Verb만 출력함.
                sentence.insert(k, senti_word[k][0])
            elif senti_word[k][1]=='Verb':
                sentence.insert(k, senti_word[k][0])



        textlist.insert(-1,sentence)

print ("Word2Vec 실행중")
testmodel = Word2Vec(textlist,window = 10,min_count=40, sg=1,iter=1000)                    # 텍스트모델 학습. 보편성을 위해 min_count 최소 10 이상 필요

print ("추출된 단어 리스트는 아래와 같습니다.")
print(list(testmodel.wv.vocab.keys()))                                                     # window는 학습 단어 / 학습 문장을 임의로 10으로 둔 것
extracted_word_list = list(testmodel.wv.vocab.keys())



print ("추출된 단어의 점수를 산정합니다.")
temp_set = list()
b = len(extracted_word_list)

for n in range (0,b):
    temp_score = 0
    temp_score_avg = 0
    temp_word = extracted_word_list[n]                                                     # word2vec의 추출 단어들 점수 설정


    for c in range (0,9):
        temp_score = temp_score + testmodel.wv.most_similar(temp_word)[c][1]                          # 추출 단어의 점수 리스트\
        print(temp_score)

    temp_score_avg = temp_score / 10
    print(temp_word)
    print(temp_score_avg)
    temp_set.append([temp_word, temp_score_avg])
    print(temp_set)


print ("텍스트 평가를 시작합니다")                                                        #이제, 평가 대상 텍스트를 찾습니다. ()
test_file = openpyxl.load_workbook('C:\\Users\\inha\\Desktop\\scored_text.xlsx')
sheet2 = test_file.active
testlist = list()

for j in range (1,27614):                                                                 #range 직접 설정해주지 않으면 오류...해결바람
        test_word = list()
        sentence = list()
        test_word = sheet2.cell(row=j, column=1).value.split("+")                                 # 엑셀 상의 텍스트 리스트로! 불러옵니당
        count=len(test_word)

        for i in range (0,count):
            test_word[i] = test_word[i].split(",")                                        # 내부 리스트도! 불러서 2차원으로 만들어냅니다. (ㅇ)

            if (len(test_word[i]) < 2) or (len(test_word[i]) > 2):
                test_word[i] = ["error","error"]
                print("에러 텍스트 삭제")





        senti_score = 0

        for k in range (0, len(test_word)):


            for temping in range (0,len(temp_set)):                                                                            #여기서 반의어/확산어/중립어 제거

                if test_word[k][0] == temp_set[temping][0]:

                    if test_word[k][0] in pre_anti_word_list:
                        test_word[k] = ["pre_anti_word", "test_word[k][0]"]  # 부정어 노드 설정

                    elif test_word[k][0] in after_anti_word_list:
                        test_word[k] = ["after_anti_word", "test_word[k][0]"]  # 부정어 노드 설정

                    elif test_word[k][0] in accent_word_list:
                        test_word[k] = ["accent_word", "test_word[k][0]"]  # 강조어 노드 설정

                    elif test_word[k][0] in trash_word_list:
                        test_word[k] = ["trash_word", "test_word[k][0]"]  # 파기어 노드 설정

                    ##노드 설정하고 이제 점수 산정합니다!

                    if (k >= 1) and (test_word[k-1][0] == "pre_anti_word"):
                        senti_score = senti_score - temp_set[temping][1]

                    elif (k+1 <= len(test_word)) and (test_word[k+1][0] == "after_anti_word"):
                        senti_score = senti_score - temp_set[temping][1]

                    elif (k >= 1) and (test_word[k-1][0] == "accent_word"):
                        senti_score = senti_score + (1.2*temp_set[temping][1])

                    elif test_word[k][0] == "trash_word":
                        senti_score = senti_score + 0

                    else:
                        senti_score = senti_score + temp_set[temping][1]

        sheet2.cell(row=j, column=2).value = senti_score

test_file.save('C:\\Users\\inha\\Desktop\\scored_text.xlsx')
print ("텍스트 평가를 완료하였습니다.")